
import os
import openpyxl

from CRUD.CRUD_tiendas import *

def actualizar_tiendas(ruta_completa):
    progreso = 0
    
    wb = openpyxl.load_workbook(ruta_completa)
    
    ws = wb['TIENDAS']
    fila = 1
    ultimafila = ws.max_row
    ultimacolumna = ws.max_column
    datos = [] #LISTA
    while fila <= ultimafila:
        columna = 1
        while columna <= ultimacolumna:
            datos.append(ws.cell(row = fila,column = columna).value)
            columna += 1
            
        
        update_tienda(datos)
        datos.clear()
        
        actual = int(fila/ultimafila *100)
        if progreso < actual:
            progreso = actual
            os.system('cls')
            print(progreso, "%")
        fila += 1
        if fila > ultimafila:
            break
    print("Completado")