import os
import openpyxl
from CRUD.CRUD_tc_rc import *

def cargar_recargo_cambio(ruta_completa):
    progreso = 0
    
    wb = openpyxl.load_workbook(ruta_completa)
    hoja = 'TC_RC'
    ws = wb[hoja]
    ultimafila = ws.max_row
    datos = []
    
    for fila in range(2,ultimafila+1):
        for columna in range(1,4):
            datos.append(ws.cell(row = fila,column = columna).value)
        
        registrar_TC_RC(datos)
        datos.clear()
        
        actual = int(fila/ultimafila *100)
        if progreso < actual:
            os.system("cls")
            progreso = actual
            print(progreso, "%")
        fila += 1
        
    print("Completado")

def actualizar_recargo_cambio(ruta_completa):
    progreso = 0
    
    wb = openpyxl.load_workbook(ruta_completa)
    hoja = 'TC_RC'
    ws = wb[hoja]
    ultimafila = ws.max_row
    datos = []
    
    for fila in range(2,ultimafila+1):
        datos.append(ws.cell(row = fila,column = 2).value)
        datos.append(ws.cell(row = fila,column = 3).value)
        datos.append(ws.cell(row = fila,column = 1).value)
        
        actualizar_TC_RC(datos)
        datos.clear()
        
        actual = int(fila/ultimafila *100)
        if progreso < actual:
            os.system("cls")
            progreso = actual
            print(progreso, "%")
        fila += 1
        
    print("Completado")