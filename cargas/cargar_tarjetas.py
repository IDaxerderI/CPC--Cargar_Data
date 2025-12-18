import openpyxl
import sys, os

from CRUD.CRUD_tiendas import serie_cod_comercio
from CRUD.CRUD_tarjetas import registar_tarjetas,registrar_culqui

from funciones.notas_temporal import crear_y_abrir_txt

def cargar_tarjetas_izipay(ruta_completa):
    print("Leyendo Archivo Excel")
    wb = openpyxl.load_workbook(ruta_completa, data_only=True)
    ws = wb['Base']
    fila = 2
    ultimafila = ws.max_row
    datos = [] #LISTA
    filas_erroneas = []
    progreso = 0
    while fila <= ultimafila:
        columnas = [2,5,6,7,8,10,11,12,13,14,15,16,17,18,19,20,22,23,24,25,26,27,28,29,30,31,32,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50]
        if ws.cell(row = fila,column = 23).value == "Abonado":
            
            for columna in columnas:
                if ws.cell(row = fila,column = columna).value != None and ws.cell(row = fila,column = columna).value != "" :
                    datos.append(ws.cell(row = fila,column = columna).value)
                else:
                    datos.append(None)
            
            fecha =ws.cell(row = fila,column = 19).value.strftime('%d/%m/%Y').replace("/","")
            hora = str(ws.cell(row = fila,column = 20).value).replace(":","")
            id = str(ws.cell(row = fila,column = 14).value)
            producto = str(ws.cell(row = fila,column = 5).value)
            if id != "None":
                serie = serie_cod_comercio(str(ws.cell(row = fila,column = 2).value))
                try:
                    id = serie + id + fecha + hora + producto
                except Exception as error:
                    print(f"fila:-{fila}")
                    print(f"fila:-{serie}")
                    print(f"fila:-{id}")
                    print(f"fila:-{fecha}")
                    print(f"fila:-{hora}")
                    print(f"Error: {error}")
                datos.insert(0,id)
                reg = registar_tarjetas(datos)
                datos.clear()
                
                if reg:
                    filas_erroneas.append(fila)
            else:
                filas_erroneas.append(fila)
                #for i in range (0,len(datos)):
                    #print(datos[i] , "-" , type(datos[i]))
            actual = int(fila/ultimafila *100)
            if progreso < actual:
                progreso = actual
                os.system('cls')
                print(progreso, "%")
            fila += 1
        else:
            fila += 1
    print("100%")
    print("Completado")
    filas_erroneas.insert(0,"FILAS ERRONEAS:")
    crear_y_abrir_txt(filas_erroneas)

def cargar_tarjetas_culqui(ruta_completa):
    print("Leyendo Archivo Excel")
    wb = openpyxl.load_workbook(ruta_completa, data_only=True)
    ws = wb['Base']
    fila = 2
    ultimafila = ws.max_row
    ultimacolumna = ws.max_column
    datos = [] #LISTA
    filas_erroneas = []
    progreso = 0
    while fila <= ultimafila:
        #columnas = [2,5,6,7,8,10,11,12,13,14,15,16,17,18,19,20,22,23,24,25,26,27,28,29,30,31,32,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50]
        columna = 1
        
        while columna <= ultimacolumna:
            if ws.cell(row = fila,column = columna).value != None and ws.cell(row = fila,column = columna).value != "" :
                datos.append(ws.cell(row = fila,column = columna).value)
            else:
                datos.append(None)
            columna += 1
        
        fecha =ws.cell(row = fila,column = 7).value.strftime('%d/%m/%Y').replace("/","")
        hora = str(ws.cell(row = fila,column = 8).value).replace(":","")
        id = str(ws.cell(row = fila,column = 24).value)
        producto = str(ws.cell(row = fila,column = 3).value)
        if id != "None":
            serie = serie_cod_comercio(str(ws.cell(row = fila,column = 35).value))
            try:
                id = serie + id + fecha + hora + producto
            except Exception as error:
                print(f"fila:-{fila}")
                print(f"fila:-{serie}")
                print(f"fila:-{id}")
                print(f"fila:-{fecha}")
                print(f"fila:-{hora}")
                print(f"Error: {error}")
            datos.insert(0,id)
            reg = registrar_culqui(datos)
            datos.clear()
            
            if reg:
                filas_erroneas.append(fila)
        else:
            filas_erroneas.append(fila)
            #for i in range (0,len(datos)):
                #print(datos[i] , "-" , type(datos[i]))
        actual = int(fila/ultimafila *100)
        if progreso < actual:
            progreso = actual
            os.system('cls')
            print(progreso, "%")
        fila += 1
    print("100%")
    print("Completado")
    filas_erroneas.insert(0,"FILAS ERRONEAS:")
    crear_y_abrir_txt(filas_erroneas)
