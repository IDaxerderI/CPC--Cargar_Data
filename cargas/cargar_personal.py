import os
import openpyxl

from CRUD.CRUD_personal import *

def matriz_match(ws):
    ultimafila = ws.max_row
    
    tiendas = []
    serie = []
    for fila in range(2,ultimafila+1):
        tiendas.append(ws.cell(row = fila,column = 2).value)
        serie.append(ws.cell(row = fila,column = 3).value)
        
        
    return tiendas, serie

def ret_serie(tiendas, series, tienda):
        indice = tiendas.index(tienda)
        serie = series[indice]
        return serie

def cargar_personal(ruta_completa):
    progreso = 0
    
    wb = openpyxl.load_workbook(ruta_completa)
    
    ws_match = wb['Match Tiendas']
    tiendas, series = matriz_match(ws_match)
    
    hoja = 'PERSONAL'
    ws = wb[hoja]
    ultimafila = ws.max_row
    ultimacolumna = ws.max_column
    
    datos = []
    
    for fila in range(2,ultimafila+1):
        for columna in range(1,ultimacolumna+1):
            if ws.cell(row = fila,column = columna).value != None and ws.cell(row = fila,column = columna).value != "" :
                datos.append(ws.cell(row = fila,column = columna).value)
            else:
                datos.append(None)
        print(f"datos -> {datos}")
        
        datos[3] = ret_serie(tiendas, series, datos[3])
        registrar_personal(datos)
        datos.clear()
        
        actual = int(fila/ultimafila *100)
        if progreso < actual:
            os.system("cls")
            progreso = actual
            print(progreso, "%")
        fila += 1
        
    print("Completado")

def actualizar_personal(ruta_completa):
    progreso = 0
    
    wb = openpyxl.load_workbook(ruta_completa)
    
    ws_match = wb['Match Tiendas']
    tiendas, series = matriz_match(ws_match)
    
    ws = wb['PERSONAL']
    fila = 2
    ultimafila = ws.max_row
    datos = [] #LISTA
    while fila <= ultimafila:
        datos.append(ws.cell(row = fila,column = 2).value)
        datos.append(ws.cell(row = fila,column = 3).value)
        datos.append(ret_serie(tiendas, series, ws.cell(row = fila,column = 4).value))
        datos.append(ws.cell(row = fila,column = 5).value)
        
        if ws.cell(row = fila,column = 6).value != "":
            #datos.append(datetime.strptime(ws.cell(row = fila,column = 5).value,"%d/%m/%Y").date())
            datos.append(ws.cell(row = fila,column = 6).value)
        else:
            datos.append(None)
        
        if ws.cell(row = fila,column = 7).value != "":
            datos.append(ws.cell(row = fila,column = 7).value)
        else:
            datos.append(None)
        
        datos.append(ws.cell(row = fila,column = 1).value)
        
        update_personal(datos)
        datos.clear()
        
        actual = int(fila/ultimafila *100)
        if progreso < actual:
            progreso = actual
            os.system('cls')
            print(progreso, "%")
        fila += 1
        if fila > ultimafila:
            break
    print("Completado")