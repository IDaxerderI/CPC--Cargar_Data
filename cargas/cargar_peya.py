import os
import openpyxl
from CRUD.CRUD_peya import *

def cargar_peya(ruta_completa):
    progreso = 0
    wb = openpyxl.load_workbook(ruta_completa)
    #RESARCIMIENTO - 35
    ws = wb['Peya 35']
    fila = 2
    ultimafila = ws.max_row
    datos = [] #LISTA
    progreso = 0
    while fila <= ultimafila:
        lista = [
                2,#ID_PEDIDO
                3,#ID RESTAURANTE
                4,#ENTREGA
                -1,#FORMA PAGO
                -1,#ESTADO PEDIDO
                8,#FECHA
                9,#HORA
                -1,#TOTAL
                16#RESARCIMIENTO
                ]
        for columna in lista:
            if columna == -1:
                datos.append(None)
            else:
                datos.append(ws.cell(row = fila,column = columna).value)
        registrar_peya(datos)
        datos.clear()
        
        actual = int(fila/ultimafila *100)
        if progreso < actual:
            os.system("cls")
            progreso = actual
            print(progreso, "%")
        fila += 1
        if fila > ultimafila:
            break
        
    print("Completado Carga 35")

    #Normal - 100
    ws = wb['Peya 100']
    fila = 2
    ultimafila = ws.max_row
    datos = [] #LISTA
    progreso = 0
    while fila <= ultimafila:
        if ws.cell(row = fila,column = 22).value == None and ws.cell(row = fila,column = 4).value != "#N/A":
            lista = [
                3,#ID_PEDIDO
                4,#ID RESTAURANTE
                0,#ENTREGA
                17,#FORMA PAGO
                -1,#ESTADO PEDIDO
                7,#FECHA
                8,#HORA
                16,#TOTAL
                -1#RESARCIMIENTO
                ]
            for columna in lista:
                if columna == -1:
                    datos.append(None)
                elif columna == 0:
                    datos.append("ENTREGADO")
                else:
                    datos.append(ws.cell(row = fila,column = columna).value)
                #MODIFICAR EL ID - EL PRIMER DATO
            registrar_peya(datos)
            datos.clear()
            
            actual = int(fila/ultimafila *100)
            if progreso < actual:
                os.system("cls")
                progreso = actual
                print("Completado Carga 35")                
                print(progreso, "%")
            fila += 1
            if fila > ultimafila:
                break
        else:
            fila += 1
            
    os.system("cls")
    print("Completado Carga 35")
    print("Completado Carga 100")
    #RESARCIMIENTO
