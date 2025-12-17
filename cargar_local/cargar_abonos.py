import openpyxl
from CRUD_local.CRUD_cuadres import *


def cargar_abonos(ruta_completa):
    wb = openpyxl.load_workbook(ruta_completa)
    
    ws = wb["Abonos"]
    ultimafila = ws.max_row
    ultimacolumna = ws.max_column
    
    datos = []
    
    for fila in range(2,ultimafila+1):
        for columna in range(1,ultimacolumna+1):
            if ws.cell(row = fila,column = columna).value != None and ws.cell(row = fila,column = columna).value != "" :
                datos.append(ws.cell(row = fila,column = columna).value)
            else:
                datos.append(None)
        
        cargar(datos)
        return 0
        datos.clear()
        
    print("Completado")