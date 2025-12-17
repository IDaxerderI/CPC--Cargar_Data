import os
import openpyxl
from CRUD_local.CRUD_anulaciones import *


def cargar_anulaciones(ruta_completa):
    wb = openpyxl.load_workbook(ruta_completa)
    
    ws = wb["Anulaciones"]
    ultimafila = ws.max_row
    ultimacolumna = ws.max_column
    
    datos = []
    
    for fila in range(2,ultimafila+1):
        for columna in range(1,ultimacolumna+1):
            if ws.cell(row = fila,column = columna).value != None and ws.cell(row = fila,column = columna).value != "" :
                datos.append(ws.cell(row = fila,column = columna).value)
            else:
                datos.append(None)
        
        cargar(datos)
        datos.clear()
        
    print("Completado")

def actualizar_anulaciones(ruta_completa):
    wb = openpyxl.load_workbook(ruta_completa)
    
    ws = wb["Anulaciones"]
    ultimafila = ws.max_row
    ultimacolumna = ws.max_column
    
    datos = []
    
    for fila in range(2,ultimafila+1):
        for columna in range(2,ultimacolumna+1):
            if ws.cell(row = fila,column = columna).value != None and ws.cell(row = fila,column = columna).value != "" :
                datos.append(ws.cell(row = fila,column = columna).value)
            else:
                datos.append(None)
        
        datos.append(ws.cell(row = fila,column = 1).value)
        actualizar(datos)
        datos.clear()
        
    print("Completado")